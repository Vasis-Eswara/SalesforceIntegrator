"""
Unit tests for excel_utils.py
"""
import os
import unittest
import pytest
from unittest.mock import patch, MagicMock, mock_open
import tempfile
import openpyxl
from excel_utils import (
    generate_object_template,
    process_excel_configuration
)

class TestExcelUtils(unittest.TestCase):
    """Test the Excel utility functions"""
    
    def setUp(self):
        """Setup test data"""
        # Create a temp directory for test files
        self.temp_dir = tempfile.TemporaryDirectory()
        
        # Sample configuration data expected from Excel
        self.sample_config = {
            'object': {
                'name': 'Custom_Object__c',
                'label': 'Custom Object',
                'pluralLabel': 'Custom Objects',
                'description': 'A test custom object'
            },
            'fields': [
                {
                    'name': 'Name__c',
                    'label': 'Name',
                    'type': 'Text',
                    'length': 255,
                    'required': True,
                    'description': 'The name field'
                },
                {
                    'name': 'Status__c',
                    'label': 'Status',
                    'type': 'Picklist',
                    'values': 'New,In Progress,Completed',
                    'required': False,
                    'description': 'The status field'
                },
                {
                    'name': 'Amount__c',
                    'label': 'Amount',
                    'type': 'Number',
                    'precision': 10,
                    'scale': 2,
                    'required': False,
                    'description': 'The amount field'
                }
            ]
        }

    def tearDown(self):
        """Clean up after tests"""
        self.temp_dir.cleanup()

    def test_generate_object_template(self):
        """Test generating an Excel template for object configuration"""
        # Create a temporary file for the template
        template_path = os.path.join(self.temp_dir.name, 'template.xlsx')
        
        # Generate the template
        result_path = generate_object_template(template_path)
        
        # Verify the file was created
        self.assertTrue(os.path.exists(result_path))
        
        # Load the workbook to verify structure
        wb = openpyxl.load_workbook(result_path)
        
        # Verify expected worksheets
        self.assertIn('Instructions', wb.sheetnames)
        self.assertIn('Object Definition', wb.sheetnames)
        self.assertIn('Fields', wb.sheetnames)
        
        # Verify object definition sheet has expected columns
        object_sheet = wb['Object Definition']
        headers = [cell.value for cell in object_sheet[1]]
        self.assertIn('Name', headers)
        self.assertIn('Label', headers)
        self.assertIn('Plural Label', headers)
        self.assertIn('Description', headers)
        
        # Verify fields sheet has expected columns
        fields_sheet = wb['Fields']
        headers = [cell.value for cell in fields_sheet[1]]
        self.assertIn('Field Name', headers)
        self.assertIn('Label', headers)
        self.assertIn('Type', headers)
        self.assertIn('Length/Precision', headers)
        self.assertIn('Scale', headers)
        self.assertIn('Required', headers)
        self.assertIn('Description', headers)
        self.assertIn('Picklist Values', headers)

    def test_process_excel_configuration_valid(self):
        """Test processing a valid Excel configuration file"""
        # Create a sample Excel file
        excel_path = os.path.join(self.temp_dir.name, 'config.xlsx')
        
        # Create a workbook with test data
        wb = openpyxl.Workbook()
        
        # Set up Object Definition sheet
        obj_sheet = wb.active
        obj_sheet.title = 'Object Definition'
        headers = ['Name', 'Label', 'Plural Label', 'Description']
        for col, header in enumerate(headers, 1):
            obj_sheet.cell(row=1, column=col, value=header)
        
        # Add object info
        obj_data = ['Custom_Object__c', 'Custom Object', 'Custom Objects', 'A test custom object']
        for col, value in enumerate(obj_data, 1):
            obj_sheet.cell(row=2, column=col, value=value)
        
        # Set up Fields sheet
        fields_sheet = wb.create_sheet('Fields')
        headers = ['Field Name', 'Label', 'Type', 'Length/Precision', 'Scale', 'Required', 'Description', 'Picklist Values']
        for col, header in enumerate(headers, 1):
            fields_sheet.cell(row=1, column=col, value=header)
        
        # Add field rows
        field_rows = [
            ['Name__c', 'Name', 'Text', 255, '', True, 'The name field', ''],
            ['Status__c', 'Status', 'Picklist', '', '', False, 'The status field', 'New,In Progress,Completed'],
            ['Amount__c', 'Amount', 'Number', 10, 2, False, 'The amount field', '']
        ]
        
        for row_idx, row_data in enumerate(field_rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                fields_sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Save the workbook
        wb.save(excel_path)
        
        # Process the configuration
        result = process_excel_configuration(excel_path)
        
        # Verify object details
        self.assertEqual(result['object']['name'], 'Custom_Object__c')
        self.assertEqual(result['object']['label'], 'Custom Object')
        self.assertEqual(result['object']['pluralLabel'], 'Custom Objects')
        self.assertEqual(result['object']['description'], 'A test custom object')
        
        # Verify fields
        self.assertEqual(len(result['fields']), 3)
        
        # Check text field properties
        name_field = result['fields'][0]
        self.assertEqual(name_field['name'], 'Name__c')
        self.assertEqual(name_field['label'], 'Name')
        self.assertEqual(name_field['type'], 'Text')
        self.assertEqual(name_field['length'], 255)
        self.assertEqual(name_field['required'], True)
        
        # Check picklist field properties
        status_field = result['fields'][1]
        self.assertEqual(status_field['name'], 'Status__c')
        self.assertEqual(status_field['type'], 'Picklist')
        self.assertEqual(status_field['values'], 'New,In Progress,Completed')
        
        # Check number field properties
        amount_field = result['fields'][2]
        self.assertEqual(amount_field['name'], 'Amount__c')
        self.assertEqual(amount_field['type'], 'Number')
        self.assertEqual(amount_field['precision'], 10)
        self.assertEqual(amount_field['scale'], 2)

    def test_process_excel_configuration_invalid_file(self):
        """Test processing an invalid Excel file"""
        # Non-existent file
        with self.assertRaises(Exception):
            process_excel_configuration('non_existent_file.xlsx')
        
        # Create an empty file
        empty_file = os.path.join(self.temp_dir.name, 'empty.txt')
        with open(empty_file, 'w') as f:
            f.write('Not an Excel file')
        
        # Should raise an exception
        with self.assertRaises(Exception):
            process_excel_configuration(empty_file)

    def test_process_excel_configuration_missing_sheets(self):
        """Test processing Excel with missing required sheets"""
        # Create a sample Excel file
        excel_path = os.path.join(self.temp_dir.name, 'missing_sheets.xlsx')
        
        # Create a workbook with only one sheet
        wb = openpyxl.Workbook()
        wb.active.title = 'Not Required'
        wb.save(excel_path)
        
        # Should raise an exception about missing required sheets
        with self.assertRaises(Exception) as context:
            process_excel_configuration(excel_path)
        
        # Verify error mentions required sheets
        self.assertIn('Required sheets not found', str(context.exception))

    def test_process_excel_configuration_empty_sheets(self):
        """Test processing Excel with empty required sheets"""
        # Create a sample Excel file
        excel_path = os.path.join(self.temp_dir.name, 'empty_sheets.xlsx')
        
        # Create a workbook with empty required sheets
        wb = openpyxl.Workbook()
        wb.active.title = 'Object Definition'
        wb.create_sheet('Fields')
        wb.save(excel_path)
        
        # Should handle empty sheets gracefully with a specific error
        with self.assertRaises(Exception) as context:
            process_excel_configuration(excel_path)
        
        # Verify error mentions missing headers or data
        self.assertIn('Missing required headers', str(context.exception).lower())

if __name__ == '__main__':
    unittest.main()